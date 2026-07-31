"""
Exporta todos los registros de la tabla event a un Excel,
resolviendo campos indexados (FKs) a sus nombres legibles.
"""
import os, sys
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'settings')
sys.path.insert(0, r'd:\PROMETEO\webapp')
import django; django.setup()
from django.db import connections
import openpyxl
from openpyxl.styles import Font, PatternFill

conn = connections['propifai']
cur = conn.cursor()
cur.execute("SELECT * FROM event ORDER BY start_time DESC")
cols = [desc[0] for desc in cur.description]
eventos = cur.fetchall()
print(f"Eventos: {len(eventos)}")

# --- Resolver referencias ---

# event_type id → name
cur.execute("SELECT id, name FROM event_type")
et_map = {r[0]: r[1] for r in cur.fetchall()}

# property id → title
cur.execute("SELECT id, title FROM property")
prop_map = {r[0]: r[1] or f"ID:{r[0]}" for r in cur.fetchall()}

# contact id → full name (first_name + last_name)
cur.execute("SELECT id, first_name, last_name FROM contact")
ct_map = {}
for r in cur.fetchall():
    name = f"{r[1] or ''} {r[2] or ''}".strip()
    ct_map[r[0]] = name or f"ID:{r[0]}"

# lead id → username
cur.execute("SELECT id, username FROM lead")
ld_map = {r[0]: r[1] or f"ID:{r[0]}" for r in cur.fetchall()}

# match id → match_status
cur.execute("SELECT id, match_status FROM match")
mt_map = {r[0]: r[1] or f"ID:{r[0]}" for r in cur.fetchall()}

# proposal id → status
cur.execute("SELECT id, status FROM proposal")
pp_map = {r[0]: r[1] or f"ID:{r[0]}" for r in cur.fetchall()}

print(f"Tipos evento: {len(et_map)}, Propiedades: {len(prop_map)}, Contactos: {len(ct_map)}")
print(f"Leads: {len(ld_map)}, Matches: {len(mt_map)}, Propuestas: {len(pp_map)}")

col_idx = {name: i for i, name in enumerate(cols)}
fmt = lambda dt: dt.strftime('%Y-%m-%d %H:%M') if dt else ''
BOOL = lambda v: 'Sí' if v else 'No'

HEADERS = [
    ('ID', 'id'), ('Código', 'code'), ('Título', 'title'),
    ('Descripción', 'description'), ('Seguimiento', 'tracing'),
    ('Fecha Inicio', 'start_time'), ('Fecha Fin', 'end_time'),
    ('Estado', 'status'), ('Activo', 'is_active'),
    ('Tipo Evento (ID)', 'event_type_id'), ('Tipo Evento', None),
    ('Agente Asignado (ID)', 'assigned_agent_id'), ('Agente Asignado', None),
    ('Contacto (ID)', 'contact_id'), ('Contacto', None),
    ('Propiedad (ID)', 'property_id'), ('Propiedad', None),
    ('Lead (ID)', 'lead_id'), ('Lead', None),
    ('Match (ID)', 'match_id'), ('Match', None),
    ('Propuesta (ID)', 'proposal_id'), ('Propuesta', None),
    ('Creado por (ID)', 'created_by_id'), ('Creado por', None),
    ('Actualizado por (ID)', 'updated_by_id'), ('Actualizado por', None),
    ('Completado', 'completed'),
    ('Creado en', 'created_at'), ('Actualizado en', 'updated_at'),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Eventos"
hf = Font(bold=True, color="FFFFFF")
hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for c, (label, _) in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=c, value=label)
    cell.font = hf; cell.fill = hfill

def v(name):
    i = col_idx.get(name)
    return None if i is None else eventos[r][i]

for r in range(len(eventos)):
    row = [v('id'), v('code'), v('title'), v('description'), v('tracing'),
           fmt(v('start_time')), fmt(v('end_time')), v('status'), BOOL(v('is_active'))]
    
    et_id = v('event_type_id'); row += [et_id, et_map.get(et_id, '')]
    aa_id = v('assigned_agent_id'); row += [aa_id, ct_map.get(aa_id, '')]
    ct_id = v('contact_id'); row += [ct_id, ct_map.get(ct_id, '')]
    pr_id = v('property_id'); row += [pr_id, prop_map.get(pr_id, '')]
    ld_id = v('lead_id'); row += [ld_id, ld_map.get(ld_id, '')]
    mt_id = v('match_id'); row += [mt_id, mt_map.get(mt_id, '')]
    pp_id = v('proposal_id'); row += [pp_id, pp_map.get(pp_id, '')]
    cb_id = v('created_by_id'); row += [cb_id, ct_map.get(cb_id, '')]
    ub_id = v('updated_by_id'); row += [ub_id, ct_map.get(ub_id, '')]
    row += [BOOL(v('completed')), fmt(v('created_at')), fmt(v('updated_at'))]
    
    for c, val in enumerate(row, 1):
        ws.cell(row=r+2, column=c, value=val)

for c, (label, _) in enumerate(HEADERS, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = max(18, len(label)+3)

out = r'd:\PROMETEO\eventos_completos.xlsx'
wb.save(out)
print(f"\n✓ {out}")
print(f"  {len(eventos)} registros | {len(HEADERS)} columnas")
print(f"  Indexados resueltos: event_type→name, property→title, contact→name,")
print(f"  lead→username, match→match_status, proposal→status")
