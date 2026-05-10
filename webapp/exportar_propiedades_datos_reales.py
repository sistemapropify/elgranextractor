#!/usr/bin/env python
"""
Exporta los DATOS REALES de la tabla properties (Propifai) a un Excel,
resolviendo todas las foreign keys a sus nombres legibles.

Ejemplo: property_type_id=1 -> "Casa", district_fk_id=1 -> "Cayma", etc.
"""

import os
import sys
import django

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'settings')

try:
    django.setup()
except Exception:
    pass

from django.db import connections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def get_nombre_legible(col_name):
    """Convierte un nombre de columna SQL a un nombre legible en español."""
    nombres = {
        'id': 'ID',
        'code': 'Código',
        'codigo_unico_propiedad': 'Código Único',
        'title': 'Título',
        'description': 'Descripción',
        'antiquity_years': 'Antigüedad (años)',
        'delivery_date': 'Fecha de Entrega',
        'price': 'Precio',
        'maintenance_fee': 'Cuota de Mantenimiento',
        'has_maintenance': '¿Tiene Mantenimiento?',
        'floors': 'Pisos',
        'bedrooms': 'Dormitorios',
        'bathrooms': 'Baños',
        'half_bathrooms': 'Medios Baños',
        'garage_spaces': 'Estacionamientos',
        'land_area': 'Área Terreno (m²)',
        'built_area': 'Área Construida (m²)',
        'front_measure': 'Medida de Frente (ml)',
        'depth_measure': 'Medida de Fondo (ml)',
        'real_address': 'Dirección Real',
        'exact_address': 'Dirección Exacta',
        'coordinates': 'Coordenadas',
        'department': 'Departamento (índice)',
        'province': 'Provincia (índice)',
        'district': 'Distrito (índice)',
        'urbanization': 'Urbanización',
        'amenities': 'Amenidades',
        'zoning': 'Zonificación',
        'created_at': 'Creado',
        'updated_at': 'Actualizado',
        'is_active': '¿Activo?',
        'is_ready_for_sale': '¿Listo para Venta?',
        'is_draft': '¿Borrador?',
        'is_project': '¿Es Proyecto?',
        'project_name': 'Nombre del Proyecto',
        'ascensor': 'Ascensor',
        'availability_status': 'Estado de Disponibilidad',
        'unit_location': 'Ubicación de Unidad',
        'parking_cost': 'Costo de Estacionamiento',
        'parking_cost_included': 'Estac. Incluido en Precio',
        'wp_last_sync': 'Última Sync WordPress',
        'wp_post_id': 'Post ID WordPress',
        'wp_slug': 'Slug WordPress',
        'source': 'Fuente',
        'source_url': 'URL Fuente',
        'source_published_at': 'Publicado en Fuente',
        'has_elevator': '¿Tiene Ascensor?',
        # FK resueltas
        'property_type_nombre': 'Tipo de Propiedad',
        'operation_type_nombre': 'Tipo de Operación',
        'condition_nombre': 'Condición',
        'currency_nombre': 'Moneda',
        'status_nombre': 'Estado',
        'forma_de_pago_nombre': 'Forma de Pago',
        'district_nombre': 'Distrito (FK)',
        'urbanization_nombre': 'Urbanización (FK)',
        'garage_type_nombre': 'Tipo de Garaje',
        'property_subtype_nombre': 'Subtipo de Propiedad',
        'water_service_nombre': 'Servicio de Agua',
        'energy_service_nombre': 'Servicio Eléctrico',
        'gas_service_nombre': 'Servicio de Gas',
        'drainage_service_nombre': 'Servicio de Desagüe',
        'built_area_unit_nombre': 'Unidad Área Construida',
        'land_area_unit_nombre': 'Unidad Área Terreno',
        'assigned_agent_nombre': 'Agente Asignado',
        'created_by_nombre': 'Creado por',
        'responsible_nombre': 'Responsable',
        'owner_nombre': 'Propietario',
    }
    return nombres.get(col_name, col_name.replace('_', ' ').title())


def main():
    output_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        'propiedades_datos_reales.xlsx'
    )

    conn = connections['propifai']

    # ============================================================
    # 1. CONSTRUIR QUERY CON TODOS LOS JOINS
    # Nombres reales de tablas obtenidos de INFORMATION_SCHEMA.TABLES
    # ============================================================
    query = """
    SELECT
        p.[id],
        p.[code],
        p.[codigo_unico_propiedad],
        p.[title],
        p.[description],
        p.[antiquity_years],
        p.[delivery_date],
        p.[price],
        p.[maintenance_fee],
        p.[has_maintenance],
        p.[floors],
        p.[bedrooms],
        p.[bathrooms],
        p.[half_bathrooms],
        p.[garage_spaces],
        p.[land_area],
        p.[built_area],
        p.[front_measure],
        p.[depth_measure],
        p.[real_address],
        p.[exact_address],
        p.[coordinates],
        p.[department],
        p.[province],
        p.[district],
        p.[urbanization],
        p.[amenities],
        p.[zoning],
        p.[created_at],
        p.[updated_at],
        p.[is_active],
        p.[is_ready_for_sale],
        p.[is_draft],
        p.[is_project],
        p.[project_name],
        p.[ascensor],
        p.[availability_status],
        p.[unit_location],
        p.[parking_cost],
        p.[parking_cost_included],
        p.[wp_last_sync],
        p.[wp_post_id],
        p.[wp_slug],
        p.[source],
        p.[source_url],
        p.[source_published_at],
        p.[has_elevator],
        -- FK resueltas a nombres legibles
        pt.[name] AS [property_type_nombre],
        ot.[name] AS [operation_type_nombre],
        pc.[name] AS [condition_nombre],
        cur.[name] AS [currency_nombre],
        pstt.[name] AS [status_nombre],
        pm.[name] AS [forma_de_pago_nombre],
        pd.[name] AS [district_nombre],
        pu.[name] AS [urbanization_nombre],
        gt.[name] AS [garage_type_nombre],
        psbt.[name] AS [property_subtype_nombre],
        ws.[name] AS [water_service_nombre],
        es.[name] AS [energy_service_nombre],
        gs.[name] AS [gas_service_nombre],
        ds.[name] AS [drainage_service_nombre],
        bau.[name] AS [built_area_unit_nombre],
        lau.[name] AS [land_area_unit_nombre],
        CONCAT(aa.[first_name], ' ', aa.[last_name]) AS [assigned_agent_nombre],
        CONCAT(cb.[first_name], ' ', cb.[last_name]) AS [created_by_nombre],
        CONCAT(r.[first_name], ' ', r.[last_name]) AS [responsible_nombre],
        CONCAT(o.[first_name], ' ', o.[last_name]) AS [owner_nombre]
    FROM [dbo].[properties] p
    LEFT JOIN [dbo].[property_types] pt ON p.[property_type_id] = pt.[id]
    LEFT JOIN [dbo].[operation_types] ot ON p.[operation_type_id] = ot.[id]
    LEFT JOIN [dbo].[property_conditions] pc ON p.[condition_id] = pc.[id]
    LEFT JOIN [dbo].[currencies] cur ON p.[currency_id] = cur.[id]
    LEFT JOIN [dbo].[property_statuses] pstt ON p.[status_id] = pstt.[id]
    LEFT JOIN [dbo].[payment_methods] pm ON p.[forma_de_pago_id] = pm.[id]
    LEFT JOIN [dbo].[properties_district] pd ON p.[district_fk_id] = pd.[id]
    LEFT JOIN [dbo].[properties_urbanization] pu ON p.[urbanization_fk_id] = pu.[id]
    LEFT JOIN [dbo].[garage_types] gt ON p.[garage_type_id] = gt.[id]
    LEFT JOIN [dbo].[property_subtypes] psbt ON p.[property_subtype_id] = psbt.[id]
    LEFT JOIN [dbo].[water_service_types] ws ON p.[water_service_id] = ws.[id]
    LEFT JOIN [dbo].[energy_service_types] es ON p.[energy_service_id] = es.[id]
    LEFT JOIN [dbo].[gas_service_types] gs ON p.[gas_service_id] = gs.[id]
    LEFT JOIN [dbo].[drainage_service_types] ds ON p.[drainage_service_id] = ds.[id]
    LEFT JOIN [dbo].[measurement_units] bau ON p.[built_area_unit_id] = bau.[id]
    LEFT JOIN [dbo].[measurement_units] lau ON p.[land_area_unit_id] = lau.[id]
    LEFT JOIN [dbo].[users] aa ON p.[assigned_agent_id] = aa.[id]
    LEFT JOIN [dbo].[users] cb ON p.[created_by_id] = cb.[id]
    LEFT JOIN [dbo].[users] r ON p.[responsible_id] = r.[id]
    LEFT JOIN [dbo].[users] o ON p.[owner_id] = o.[id]
    ORDER BY p.[id]
    """

    print("Ejecutando consulta con JOINs a todas las tablas referenciadas...")
    sys.stdout.flush()
    with conn.cursor() as cursor:
        cursor.execute(query)
        rows = cursor.fetchall()
        col_names = [desc[0] for desc in cursor.description]

    total_rows = len(rows)
    total_cols = len(col_names)
    print(f"  -> {total_rows} propiedades encontradas")
    print(f"  -> {total_cols} columnas (incluyendo FK resueltas)")
    sys.stdout.flush()

    # ============================================================
    # 2. CREAR EXCEL
    # ============================================================
    wb = Workbook()

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    data_font = Font(name='Calibri', size=9)
    data_align = Alignment(vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    even_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
    odd_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # ============================================================
    # HOJA 1: DATOS DE PROPIEDADES
    # ============================================================
    ws = wb.active
    ws.title = 'Propiedades (Datos Reales)'

    for col_idx, col_name in enumerate(col_names, 1):
        header_text = get_nombre_legible(col_name)
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        row_fill = even_fill if row_idx % 2 == 0 else odd_fill
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if value is None:
                cell.value = ''
            elif isinstance(value, bool):
                cell.value = 'Sí' if value else 'No'
                cell.alignment = center_align
            elif isinstance(value, float):
                cell.value = round(value, 2)
                cell.alignment = center_align
            elif isinstance(value, int):
                cell.value = value
                cell.alignment = center_align
            else:
                cell.value = str(value)

            cell.font = data_font
            cell.border = thin_border
            cell.fill = row_fill

            if cell.alignment == Alignment():
                cell.alignment = data_align

    # Ajustar ancho de columnas
    for col_idx, col_name in enumerate(col_names, 1):
        header_len = len(get_nombre_legible(col_name))
        sample_len = header_len
        for row_idx in range(2, min(5, total_rows + 2)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                val_str = str(val)
                sample_len = max(sample_len, min(len(val_str), 60))
        width = min(max(sample_len + 2, 8), 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = 'C2'
    ws.auto_filter.ref = f"A1:{get_column_letter(total_cols)}{total_rows + 1}"

    # ============================================================
    # HOJA 2: RESUMEN
    # ============================================================
    ws2 = wb.create_sheet('Resumen')

    title_font = Font(name='Calibri', bold=True, size=14, color='2F5496')
    subtitle_font = Font(name='Calibri', bold=True, size=11, color='333333')
    text_font = Font(name='Calibri', size=10)

    ws2.cell(row=1, column=1, value='Exportación de Datos Reales - Propiedades Propifai').font = title_font
    ws2.cell(row=2, column=1, value=f'Total de propiedades: {total_rows}').font = subtitle_font
    ws2.cell(row=3, column=1, value=f'Total de columnas: {total_cols}').font = subtitle_font
    ws2.cell(row=4, column=1, value=f'Columnas FK resueltas a nombres: 20').font = subtitle_font

    ws2.cell(row=6, column=1, value='Foreign Keys Resueltas:').font = subtitle_font
    fk_list = [
        ('property_type_nombre', 'property_types', 'Tipo de propiedad (Casa, Departamento, etc.)'),
        ('operation_type_nombre', 'operation_types', 'Tipo de operación (Venta, Alquiler)'),
        ('condition_nombre', 'property_conditions', 'Condición (Nueva, Usada, En planos)'),
        ('currency_nombre', 'currencies', 'Moneda (PEN, USD)'),
        ('status_nombre', 'property_statuses', 'Estado (Disponible, Vendida, etc.)'),
        ('forma_de_pago_nombre', 'payment_methods', 'Forma de pago (Contado, Crédito)'),
        ('district_nombre', 'properties_district', 'Distrito (Cayma, Yanahuara, etc.)'),
        ('urbanization_nombre', 'properties_urbanization', 'Urbanización'),
        ('garage_type_nombre', 'garage_types', 'Tipo de garaje'),
        ('property_subtype_nombre', 'property_subtypes', 'Subtipo de propiedad'),
        ('water_service_nombre', 'water_service_types', 'Servicio de agua'),
        ('energy_service_nombre', 'energy_service_types', 'Servicio eléctrico'),
        ('gas_service_nombre', 'gas_service_types', 'Servicio de gas'),
        ('drainage_service_nombre', 'drainage_service_types', 'Servicio de desagüe'),
        ('built_area_unit_nombre', 'measurement_units', 'Unidad de área construida'),
        ('land_area_unit_nombre', 'measurement_units', 'Unidad de área de terreno'),
        ('assigned_agent_nombre', 'users', 'Agente asignado'),
        ('created_by_nombre', 'users', 'Creado por'),
        ('responsible_nombre', 'users', 'Responsable'),
        ('owner_nombre', 'users', 'Propietario'),
    ]

    for i, (col, table, desc) in enumerate(fk_list, 7):
        ws2.cell(row=i, column=1, value=col).font = Font(name='Calibri', size=10, bold=True)
        ws2.cell(row=i, column=2, value=f'← {table}').font = text_font
        ws2.cell(row=i, column=3, value=desc).font = text_font

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 55

    # ============================================================
    # GUARDAR
    # ============================================================
    wb.save(output_path)
    print(f"\nExcel generado exitosamente: {output_path}")
    print(f"  - Hoja 1: Propiedades (Datos Reales) - {total_rows} registros x {total_cols} columnas")
    print(f"  - Hoja 2: Resumen - FK resueltas y metadatos")
    print(f"\nLas foreign keys han sido resueltas a sus nombres legibles:")
    print(f"  - property_type_id -> 'Casa', 'Departamento', etc.")
    print(f"  - district_fk_id -> 'Cayma', 'Yanahuara', etc.")
    print(f"  - currency_id -> 'PEN', 'USD'")
    print(f"  - Y 17 relaciones más...")
    sys.stdout.flush()


if __name__ == '__main__':
    main()
