"""
Exporta todas las propiedades de dbpropify_be a un Excel con todos los campos
y FKs resueltas (nombres en lugar de IDs).

Uso: python manage.py exportar_propiedades_completo
"""

import os
from django.db import connections
from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


FK_RELATIONS = {
    'district_id': ('district', 'name', 'Distrito'),
    'property_type_id': ('property_type', 'name', 'Tipo Propiedad'),
    'operation_type_id': ('operation_type', 'name', 'Tipo Operación'),
    'property_condition_id': ('property_condition', 'name', 'Condición'),
    'property_status_id': ('property_status', 'name', 'Estado'),
    'currency_id': ('currency', 'name', 'Moneda'),
    'payment_method_id': ('payment_method', 'name', 'Forma Pago'),
    'property_subtype_id': ('property_subtype', 'name', 'Subtipo'),
    'urbanization_id': ('urbanization', 'name', 'Urbanización'),
    'contact_id': ('contact', 'CONCAT(first_name, \' \', last_name)', 'Contacto'),
    'responsible_id': ('[user]', 'CONCAT(first_name, \' \', last_name)', 'Responsable'),
    'created_by_id': ('[user]', 'CONCAT(first_name, \' \', last_name)', 'Creado por'),
    'updated_by_id': ('[user]', 'CONCAT(first_name, \' \', last_name)', 'Actualizado por'),
    'parent_project_id': ('property', 'title', 'Proyecto Padre'),
}

COLUMN_LABELS = {
    'id': 'ID',
    'code': 'Código',
    'uuid': 'UUID',
    'title': 'Título',
    'description': 'Descripción',
    'price': 'Precio',
    'maintenance_fee': 'Cuota Mantenimiento',
    'map_address': 'Dirección Mapa',
    'display_address': 'Dirección Mostrar',
    'latitude': 'Latitud',
    'longitude': 'Longitud',
    'registry_number': 'Nro. Registro',
    'is_project': '¿Es Proyecto?',
    'is_visible': '¿Visible?',
    'project_name': 'Nombre Proyecto',
    'video_url': 'URL Video',
    'wp_post_id': 'Post ID WP',
    'wp_slug': 'Slug WP',
    'wp_last_sync': 'Última Sync WP',
    'created_at': 'Creado',
    'updated_at': 'Actualizado',
}


class Command(BaseCommand):
    help = 'Exporta todas las propiedades de dbpropify_be a Excel con FKs resueltas'

    def handle(self, *args, **options):
        output_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))),
            'propiedades_completas.xlsx'
        )

        conn = connections['propifai']
        cursor = conn.cursor()

        # 1. Obtener columnas de property
        cursor.execute("""
            SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS 
            WHERE TABLE_NAME = 'property' AND TABLE_SCHEMA = 'dbo'
            ORDER BY ORDINAL_POSITION
        """)
        property_columns = [r[0] for r in cursor.fetchall()]

        self.stdout.write(f"Columnas en property: {len(property_columns)}")

        # 2. Obtener columnas de property_specs
        cursor.execute("""
            SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS 
            WHERE TABLE_NAME = 'property_specs' AND TABLE_SCHEMA = 'dbo'
            ORDER BY ORDINAL_POSITION
        """)
        specs_columns = [r[0] for r in cursor.fetchall()]
        # Excluir property_id (es el FK a property) y id
        specs_columns = [c for c in specs_columns if c not in ('id', 'property_id')]

        self.stdout.write(f"Columnas en property_specs (sin FK): {len(specs_columns)}")

        # 3. Construir query
        select_parts = [f"p.[{col}]" for col in property_columns]
        
        for spec_col in specs_columns:
            select_parts.append(f"ps.[{spec_col}] AS [spec_{spec_col}]")

        for fk_field, (ref_table, ref_display, label) in FK_RELATIONS.items():
            alias = fk_field.replace('_id', '_nombre')
            select_parts.append(
                f"(SELECT {ref_display} FROM {ref_table} WHERE id = p.{fk_field}) AS [{alias}]"
            )

        joins = "LEFT JOIN property_specs ps ON ps.property_id = p.id"

        query = f"""
            SELECT {', '.join(select_parts)}
            FROM [dbo].[property] p
            {joins}
            ORDER BY p.id
        """

        self.stdout.write("Ejecutando query...")
        cursor.execute(query)
        rows = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]
        self.stdout.write(f"Propiedades encontradas: {len(rows)}")

        # 4. Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Propiedades Completas"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="0d1117", end_color="0d1117", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        header_labels = []
        for col in columns:
            base_name = col.replace('spec_', '')
            label = COLUMN_LABELS.get(base_name, base_name.replace('_', ' ').title())
            if col.startswith('spec_'):
                label = f"Spec: {label}"
            if col.endswith('_nombre'):
                fk_field = col.replace('_nombre', '_id')
                if fk_field in FK_RELATIONS:
                    label = FK_RELATIONS[fk_field][2]
            header_labels.append(label)

        for col_idx, label in enumerate(header_labels, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Datos
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if value is None:
                    cell.value = ""
                elif isinstance(value, bool):
                    cell.value = "Sí" if value else "No"
                else:
                    cell.value = str(value)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # Ajustar ancho de columnas
        for col_idx in range(1, len(columns) + 1):
            max_length = len(str(header_labels[col_idx - 1]))
            for row_idx in range(2, min(len(rows) + 2, 50)):  # Sample first 50 rows
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_length = max(max_length, min(len(cell_value), 60))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        # Congelar primera fila
        ws.freeze_panes = 'A2'

        wb.save(output_path)
        self.stdout.write(self.style.SUCCESS(
            f"\n✅ Excel creado: {output_path}"
        ))
        self.stdout.write(f"   Total propiedades: {len(rows)}")
        self.stdout.write(f"   Total columnas: {len(columns)}")
