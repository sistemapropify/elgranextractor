from django.views.generic import ListView, DetailView, RedirectView, TemplateView, View
from .models import ZonaCalle, ConfiguracionCalidad
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.shortcuts import render
from django.http import JsonResponse
from django.utils import timezone
from django.db import models
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from datetime import timedelta, datetime
import json
from .models import (
    Requerimiento,
    FuenteChoices,
    CondicionChoices,
    TipoPropiedadChoices,
    MonedaChoices,
    FormaPagoChoices,
    TernarioChoices,
    TipoOriginalChoices,
    ConfiguracionCalidad,
)
from .analytics import (
    obtener_requerimientos_por_mes,
    calcular_crecimiento_porcentual,
    obtener_distritos_por_mes,
    obtener_tipos_propiedad_por_mes,
    obtener_presupuesto_por_mes,
    obtener_caracteristicas_demandadas,
    detectar_picos_y_valles,
    calcular_tendencia
)
from .tasks import generar_analisis_temporal, obtener_progreso_tarea


class ListaRequerimientosView(ListView):
    model = Requerimiento
    template_name = 'requerimientos/lista.html'
    paginate_by = 25
    context_object_name = 'requerimientos'
    ordering = ['-fecha', '-hora']

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # ── Filtros ─────────────────────────────
        q = self.request.GET.get('q', '').strip()
        fuente = self.request.GET.get('fuente', '').strip()
        condicion = self.request.GET.get('condicion', '').strip()
        tipo_propiedad = self.request.GET.get('tipo_propiedad', '').strip()
        distrito = self.request.GET.get('distrito', '').strip()
        agente = self.request.GET.get('agente', '').strip()
        moneda = self.request.GET.get('moneda', '').strip()
        forma_pago = self.request.GET.get('forma_pago', '').strip()
        cochera = self.request.GET.get('cochera', '').strip()
        ascensor = self.request.GET.get('ascensor', '').strip()
        amueblado = self.request.GET.get('amueblado', '').strip()
        tipo_original = self.request.GET.get('tipo_original', '').strip()
        verificado = self.request.GET.get('verificado', '').strip()
        quality_nivel = self.request.GET.get('quality_nivel', '').strip()
        presupuesto_min = self.request.GET.get('presupuesto_min', '').strip()
        presupuesto_max = self.request.GET.get('presupuesto_max', '').strip()
        hab_min = self.request.GET.get('hab_min', '').strip()
        area_min = self.request.GET.get('area_min', '').strip()
        fecha_desde = self.request.GET.get('fecha_desde', '').strip()
        fecha_hasta = self.request.GET.get('fecha_hasta', '').strip()
        
        # Búsqueda general (texto completo)
        if q:
            queryset = queryset.filter(
                models.Q(requerimiento__icontains=q) |
                models.Q(agente__icontains=q) |
                models.Q(distritos__icontains=q) |
                models.Q(fuente__icontains=q)
            )
        
        # Filtros exactos
        if fuente:
            queryset = queryset.filter(fuente=fuente)
        if condicion:
            queryset = queryset.filter(condicion=condicion)
        if tipo_propiedad:
            queryset = queryset.filter(tipo_propiedad=tipo_propiedad)
        if distrito:
            queryset = queryset.filter(distritos__icontains=distrito)
        if agente:
            queryset = queryset.filter(agente__icontains=agente)
        if moneda:
            queryset = queryset.filter(presupuesto_moneda=moneda)
        if forma_pago:
            queryset = queryset.filter(presupuesto_forma_pago=forma_pago)
        if cochera:
            queryset = queryset.filter(cochera=cochera)
        if ascensor:
            queryset = queryset.filter(ascensor=ascensor)
        if amueblado:
            queryset = queryset.filter(amueblado=amueblado)
        if tipo_original:
            queryset = queryset.filter(tipo_original=tipo_original)
        if verificado == 'si':
            queryset = queryset.filter(verificado=True)
        elif verificado == 'no':
            queryset = queryset.filter(verificado=False)
        if quality_nivel:
            queryset = queryset.filter(quality_nivel=quality_nivel)
        
        # Rangos numéricos
        if presupuesto_min:
            try:
                queryset = queryset.filter(presupuesto_monto__gte=float(presupuesto_min))
            except (ValueError, TypeError):
                pass
        if presupuesto_max:
            try:
                queryset = queryset.filter(presupuesto_monto__lte=float(presupuesto_max))
            except (ValueError, TypeError):
                pass
        if hab_min:
            try:
                queryset = queryset.filter(habitaciones__gte=int(hab_min))
            except (ValueError, TypeError):
                pass
        if area_min:
            try:
                queryset = queryset.filter(area_m2__gte=int(area_min))
            except (ValueError, TypeError):
                pass
        
        # Rango de fechas
        if fecha_desde:
            try:
                queryset = queryset.filter(fecha__gte=datetime.strptime(fecha_desde, '%Y-%m-%d').date())
            except (ValueError, TypeError):
                pass
        if fecha_hasta:
            try:
                queryset = queryset.filter(fecha__lte=datetime.strptime(fecha_hasta, '%Y-%m-%d').date())
            except (ValueError, TypeError):
                pass
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener fuentes reales desde la BD (valores distintos del campo 'fuente')
        fuentes_reales = (
            Requerimiento.objects.values_list('fuente', flat=True)
            .exclude(fuente='')
            .exclude(fuente__isnull=True)
            .distinct()
            .order_by('fuente')
        )
        
        # Pasar filtros actuales al template
        context['current_filters'] = {k: v for k, v in self.request.GET.items() if v}
        context['q'] = self.request.GET.get('q', '')
        context['fuentes'] = [(f, f) for f in fuentes_reales]
        context['condiciones'] = CondicionChoices.choices
        context['tipos_propiedad'] = TipoPropiedadChoices.choices
        context['monedas'] = MonedaChoices.choices
        context['formas_pago'] = FormaPagoChoices.choices
        context['ternario_opts'] = TernarioChoices.choices
        context['tipos_originales'] = TipoOriginalChoices.choices
        context['verificados_count'] = Requerimiento.objects.filter(verificado=True).count()
        context['no_verificados_count'] = Requerimiento.objects.filter(verificado=False).count()
        context['total_count'] = Requerimiento.objects.count()
        
        return context


class DetalleRequerimientoView(DetailView):
    model = Requerimiento
    template_name = 'requerimientos/detalle.html'
    context_object_name = 'requerimiento'


class SubirExcelView(RedirectView):
    """Redirige a la subida de Excel de la app ingestas."""
    pattern_name = 'ingestas:subir_excel'


# ─────────────────────────────────────────────
#  VISTAS DE ANÁLISIS TEMPORAL
# ─────────────────────────────────────────────

class DashboardAnalisisTemporalView(TemplateView):
    """Vista principal del dashboard de análisis temporal."""
    template_name = 'requerimientos/dashboard_analisis.html'
    
    def post(self, request, *args, **kwargs):
        """Toggle verificado vía POST (AJAX)."""
        import json
        try:
            data = json.loads(request.body)
            pk = data.get('pk')
            verificado = data.get('verificado')
        except (json.JSONDecodeError, TypeError):
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        if pk is None or verificado is None:
            return JsonResponse({'error': 'Faltan pk o verificado'}, status=400)

        req = get_object_or_404(Requerimiento, pk=pk)
        req.verificado = bool(verificado)
        req.save(update_fields=['verificado'])
        return JsonResponse({'ok': True, 'verificado': req.verificado})


class ToggleVerificadoView(View):
    """Vista para toggle rápido de verificado vía POST."""
    def post(self, request, *args, **kwargs):
        import json
        try:
            data = json.loads(request.body)
            pk = data.get('pk')
            verificado = data.get('verificado')
        except (json.JSONDecodeError, TypeError):
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        if pk is None:
            return JsonResponse({'error': 'Falta pk'}, status=400)

        req = get_object_or_404(Requerimiento, pk=pk)
        if verificado is not None:
            req.verificado = bool(verificado)
        else:
            req.verificado = not req.verificado  # toggle
        req.save(update_fields=['verificado'])
        return JsonResponse({'ok': True, 'verificado': req.verificado})


class EditarRequerimientoView(View):
    """Vista para editar un requerimiento vía POST (AJAX)."""
    CAMPOS_EDITABLES = [
        'verificado', 'fuente', 'agente', 'agente_telefono',
        'condicion', 'tipo_propiedad', 'distritos',
        'urbanizacion', 'zona',
        'presupuesto_monto', 'presupuesto_moneda', 'presupuesto_forma_pago',
        'habitaciones', 'banos', 'cochera', 'ascensor', 'amueblado',
        'area_m2', 'piso_preferencia', 'caracteristicas_extra',
        'tipo_original', 'requerimiento',
    ]

    def post(self, request, *args, **kwargs):
        import json
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, TypeError):
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        pk = data.get('pk')
        if not pk:
            return JsonResponse({'error': 'Falta pk'}, status=400)

        req = get_object_or_404(Requerimiento, pk=pk)
        campos_actualizados = []

        for campo in self.CAMPOS_EDITABLES:
            if campo in data:
                valor = data[campo]
                # Guardia de seguridad para campos numéricos:
                # NO sobrescribir un valor existente con None/vacío
                if campo in ('presupuesto_monto', 'area_m2', 'habitaciones', 'banos'):
                    valor_actual = getattr(req, campo)
                    if valor == '' or valor is None:
                        # Si ya hay un valor en DB, preservarlo
                        if valor_actual is not None:
                            continue
                        # Si no hay valor en DB, mantenerlo como None
                        continue
                    else:
                        try:
                            valor_convertido = float(valor) if campo == 'presupuesto_monto' else int(valor)
                            setattr(req, campo, valor_convertido)
                        except (ValueError, TypeError):
                            return JsonResponse({'error': f'Valor inválido para {campo}: {valor}'}, status=400)
                else:
                    setattr(req, campo, valor)
                campos_actualizados.append(campo)

        # ── Sincronizar ZonaCalle con los tags del campo 'zona' ──
        if 'zona' in data and data['zona']:
            tags = [t.strip() for t in data['zona'].split(',') if t.strip()]
            for tag in tags:
                obj, created = ZonaCalle.objects.get_or_create(
                    nombre__iexact=tag,
                    defaults={'nombre': tag, 'veces_usado': 1}
                )
                if not created:
                    # Si existe pero con diferente capitalización, actualizar
                    if obj.nombre != tag:
                        obj.nombre = tag
                    ZonaCalle.objects.filter(pk=obj.pk).update(veces_usado=models.F('veces_usado') + 1)

        if campos_actualizados:
            req.save(update_fields=campos_actualizados)

        return JsonResponse({
            'ok': True,
            'pk': req.pk,
            'verificado': req.verificado,
            'actualizado_en': req.actualizado_en.strftime('%d/%m/%Y %H:%M') if req.actualizado_en else '',
        })


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Obtener filtros de la URL
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')
        condicion = self.request.GET.get('condicion')
        tipo_propiedad = self.request.GET.get('tipo_propiedad')
        distrito = self.request.GET.get('distrito')
        fuente = self.request.GET.get('fuente')
        
        # Convertir fechas
        if fecha_inicio:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        if fecha_fin:
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        
        filtros = {
            'condicion': condicion,
            'tipo_propiedad': tipo_propiedad,
            'distrito': distrito,
            'fuente': fuente,
        }
        
        # Obtener datos para el dashboard
        context['filtros'] = filtros
        context['fecha_inicio'] = fecha_inicio
        context['fecha_fin'] = fecha_fin
        
        # Opciones para filtros
        context['fuentes'] = FuenteChoices.choices
        context['condiciones'] = CondicionChoices.choices
        context['tipos_propiedad'] = TipoPropiedadChoices.choices
        
        return context


class ApiAnalisisTemporalView(TemplateView):
    """API que retorna datos JSON para el dashboard (modo síncrono o asíncrono)."""
    
    def get(self, request, *args, **kwargs):
        # Verificar si se solicita modo asíncrono
        async_mode = request.GET.get('async', 'false').lower() == 'true'
        
        if async_mode:
            return self._iniciar_analisis_asincrono(request)
        else:
            return self._obtener_analisis_sincrono(request)
    
    def _obtener_analisis_sincrono(self, request):
        """Retorna análisis inmediato (para datasets pequeños)."""
        try:
            # Obtener filtros
            fecha_inicio = request.GET.get('fecha_inicio')
            fecha_fin = request.GET.get('fecha_fin')
            condicion = request.GET.get('condicion')
            tipo_propiedad = request.GET.get('tipo_propiedad')
            distrito = request.GET.get('distrito')
            fuente = request.GET.get('fuente')
            
            # Convertir fechas
            if fecha_inicio:
                fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            if fecha_fin:
                fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            
            filtros = {
                'condicion': condicion,
                'tipo_propiedad': tipo_propiedad,
                'distrito': distrito,
                'fuente': fuente,
            }
            
            # Obtener todos los datos con manejo de errores individual
            try:
                datos_mes = list(obtener_requerimientos_por_mes(fecha_inicio, fecha_fin, filtros))
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error en obtener_requerimientos_por_mes: {e}")
                datos_mes = []
            
            try:
                distritos_mes = obtener_distritos_por_mes(fecha_inicio, fecha_fin)
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error en obtener_distritos_por_mes: {e}")
                distritos_mes = {'distritos': [], 'data': {}, 'meses': []}
            
            try:
                tipos_mes = obtener_tipos_propiedad_por_mes(fecha_inicio, fecha_fin)
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error en obtener_tipos_propiedad_por_mes: {e}")
                tipos_mes = {'tipos': [], 'data': {}, 'meses': []}
            
            try:
                presupuesto_mes = obtener_presupuesto_por_mes(fecha_inicio, fecha_fin)
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error en obtener_presupuesto_por_mes: {e}")
                presupuesto_mes = []
            
            try:
                caracteristicas_mes = obtener_caracteristicas_demandadas(fecha_inicio, fecha_fin)
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error en obtener_caracteristicas_demandadas: {e}")
                caracteristicas_mes = {}
            
            # Calcular crecimiento y tendencias
            totales = [item['total'] for item in datos_mes] if datos_mes else []
            try:
                crecimiento = calcular_crecimiento_porcentual(totales)
            except Exception as e:
                crecimiento = []
            
            try:
                picos, valles = detectar_picos_y_valles(totales)
            except Exception as e:
                picos, valles = [], []
            
            try:
                tendencia = calcular_tendencia(totales)
            except Exception as e:
                tendencia = 'estable'
            
            # Generar insights
            try:
                insights = self._generar_insights(
                    datos_mes, distritos_mes, tipos_mes, presupuesto_mes
                )
            except Exception as e:
                insights = []
            
            # Preparar respuesta
            response_data = {
                'success': True,
                'mode': 'sync',
                'datos_mes': datos_mes,
                'distritos_mes': distritos_mes,
                'tipos_mes': tipos_mes,
                'presupuesto_mes': presupuesto_mes,
                'caracteristicas_mes': caracteristicas_mes,
                'metricas': {
                    'totales': totales,
                    'crecimiento': crecimiento,
                    'picos': picos,
                    'valles': valles,
                    'tendencia': tendencia,
                },
                'insights': insights,
            }
            
            return JsonResponse(response_data, safe=False)
            
        except Exception as e:
            # Log del error para diagnóstico
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error en _obtener_analisis_sincrono: {e}", exc_info=True)
            
            # Intentar devolver datos de ejemplo si la base de datos está vacía
            try:
                # Verificar si hay datos en la base de datos
                from .models import Requerimiento
                if Requerimiento.objects.exists():
                    # Hay datos pero hay un error en el procesamiento
                    return JsonResponse({
                        'success': False,
                        'error': str(e),
                        'mode': 'sync',
                        'message': 'Error procesando datos existentes'
                    }, status=500)
                else:
                    # Base de datos vacía, devolver datos de ejemplo para demostración
                    return self._devolver_datos_ejemplo()
            except Exception as inner_e:
                # Si falla la verificación, devolver error genérico
                return JsonResponse({
                    'success': False,
                    'error': f"{str(e)} (adicional: {str(inner_e)})",
                    'mode': 'sync',
                    'message': 'Error crítico en el análisis'
                }, status=500)
    
    def _iniciar_analisis_asincrono(self, request):
        """Inicia análisis asíncrono y retorna ID de tarea."""
        try:
            # Obtener filtros
            fecha_inicio = request.GET.get('fecha_inicio')
            fecha_fin = request.GET.get('fecha_fin')
            condicion = request.GET.get('condicion')
            tipo_propiedad = request.GET.get('tipo_propiedad')
            distrito = request.GET.get('distrito')
            fuente = request.GET.get('fuente')
            
            filtros = {
                'condicion': condicion,
                'tipo_propiedad': tipo_propiedad,
                'distrito': distrito,
                'fuente': fuente,
            }
            
            # Lanzar tarea asíncrona
            task = generar_analisis_temporal.delay(
                filtros=filtros,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin
            )
            
            return JsonResponse({
                'success': True,
                'mode': 'async',
                'task_id': task.id,
                'status_url': f'/requerimientos/api/analisis-progreso/{task.id}/',
                'message': 'Análisis iniciado en segundo plano'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e),
                'mode': 'async'
            }, status=500)
    
    def _generar_insights(self, datos_mes, distritos_mes, tipos_mes, presupuesto_mes):
        """Genera insights automáticos en lenguaje natural."""
        insights = []
        
        if not datos_mes:
            return insights
        
        # Insight 1: Mes con más demanda
        max_mes = max(datos_mes, key=lambda x: x['total'])
        min_mes = min(datos_mes, key=lambda x: x['total'])
        
        insights.append({
            'icono': '🔥',
            'titulo': f'Mes pico: {max_mes["mes"].strftime("%B %Y")}',
            'descripcion': f'{max_mes["total"]} requerimientos (+{max_mes["total"] - min_mes["total"]} vs mes más bajo)',
            'tipo': 'destacado'
        })
        
        # Insight 2: Crecimiento total
        if len(datos_mes) >= 2:
            primer_mes = datos_mes[0]['total']
            ultimo_mes = datos_mes[-1]['total']
            if primer_mes > 0:
                crecimiento_total = ((ultimo_mes - primer_mes) / primer_mes) * 100
                tendencia = '📈' if crecimiento_total > 0 else '📉'
                insights.append({
                    'icono': tendencia,
                    'titulo': f'Crecimiento total: {crecimiento_total:.1f}%',
                    'descripcion': f'De {primer_mes} a {ultimo_mes} requerimientos',
                    'tipo': 'tendencia'
                })
        
        # Insight 3: Distrito líder
        if distritos_mes['distritos']:
            # Calcular total por distrito
            distrito_totales = {}
            for distrito in distritos_mes['distritos']:
                total = sum(distritos_mes['data'][distrito].values())
                distrito_totales[distrito] = total
            
            top_distrito = max(distrito_totales.items(), key=lambda x: x[1])
            insights.append({
                'icono': '🏆',
                'titulo': f'Distrito líder: {top_distrito[0]}',
                'descripcion': f'{top_distrito[1]} requerimientos en el período',
                'tipo': 'liderazgo'
            })
        
        # Insight 4: Tipo de propiedad más buscado
        if tipos_mes['tipos']:
            tipo_totales = {}
            for tipo in tipos_mes['tipos']:
                total = sum(tipos_mes['data'][tipo])
                tipo_totales[tipo] = total
            
            top_tipo = max(tipo_totales.items(), key=lambda x: x[1])
            tipo_display = dict(TipoPropiedadChoices.choices).get(top_tipo[0], top_tipo[0])
            insights.append({
                'icono': '🏠',
                'titulo': f'Tipo más buscado: {tipo_display}',
                'descripcion': f'{top_tipo[1]} requerimientos ({top_tipo[1]/sum(tipo_totales.values())*100:.1f}%)',
                'tipo': 'preferencia'
            })
        
        return insights

    def _devolver_datos_ejemplo(self):
        """Devuelve datos de ejemplo para demostración cuando la base de datos está vacía."""
        from datetime import timedelta
        import random
        
        # Generar datos de ejemplo para los últimos 6 meses
        meses = []
        fecha_actual = timezone.now().date()
        for i in range(6):
            mes = fecha_actual.replace(day=1) - timedelta(days=i*30)
            meses.append({
                'mes': mes,
                'total': random.randint(10, 50),
                'compra': random.randint(5, 25),
                'alquiler': random.randint(5, 25),
                'departamento': random.randint(3, 20),
                'casa': random.randint(2, 15),
                'terreno': random.randint(1, 10),
                'presupuesto_promedio': random.uniform(50000, 200000),
                'presupuesto_mediano': None,
                'cochera_si': random.randint(0, 10),
                'ascensor_si': random.randint(0, 5),
                'amueblado_si': random.randint(0, 8),
            })
        
        # Datos de ejemplo para distritos
        distritos_ejemplo = ['Cayma', 'Yanahuara', 'Cerro Colorado', 'Sachaca', 'Hunter']
        distritos_data = {}
        for distrito in distritos_ejemplo:
            distritos_data[distrito] = {mes['mes'].strftime('%Y-%m'): random.randint(1, 10) for mes in meses}
        
        # Datos de ejemplo para tipos de propiedad
        tipos_ejemplo = ['departamento', 'casa', 'terreno', 'oficina', 'local_comercial']
        tipos_data = {}
        for tipo in tipos_ejemplo:
            tipos_data[tipo] = [random.randint(1, 15) for _ in range(len(meses))]
        
        # Datos de ejemplo para presupuesto
        presupuesto_data = [{
            'mes': mes['mes'],
            'presupuesto_promedio': mes['presupuesto_promedio'],
            'presupuesto_mediano': None,
            'cantidad': mes['total']
        } for mes in meses]
        
        # Datos de ejemplo para características
        caracteristicas_data = {
            'cochera': random.randint(20, 80),
            'ascensor': random.randint(10, 50),
            'amueblado': random.randint(15, 60),
            'habitaciones_3': random.randint(5, 40),
            'banos_2': random.randint(10, 50),
        }
        
        # Calcular métricas
        totales = [mes['total'] for mes in meses]
        crecimiento = [None] + [random.uniform(-20, 30) for _ in range(len(meses)-1)]
        picos = [i for i, val in enumerate(totales) if val == max(totales)]
        valles = [i for i, val in enumerate(totales) if val == min(totales)]
        tendencia = 'creciente' if totales[-1] > totales[0] else 'decreciente'
        
        # Insights de ejemplo
        insights = [
            {
                'icono': '🔥',
                'titulo': 'Mes pico: ' + meses[picos[0]]['mes'].strftime('%B %Y'),
                'descripcion': f'{totales[picos[0]]} requerimientos',
                'tipo': 'destacado'
            },
            {
                'icono': '📈',
                'titulo': f'Crecimiento total: {((totales[-1] - totales[0]) / totales[0] * 100):.1f}%',
                'descripcion': f'De {totales[0]} a {totales[-1]} requerimientos',
                'tipo': 'tendencia'
            },
            {
                'icono': '🏆',
                'titulo': f'Distrito líder: {distritos_ejemplo[0]}',
                'descripcion': f'{sum(distritos_data[distritos_ejemplo[0]].values())} requerimientos',
                'tipo': 'liderazgo'
            },
            {
                'icono': '🏠',
                'titulo': f'Tipo más buscado: Departamento',
                'descripcion': f'{sum(tipos_data["departamento"])} requerimientos',
                'tipo': 'preferencia'
            },
        ]
        
        response_data = {
            'success': True,
            'mode': 'sync',
            'datos_mes': meses,
            'distritos_mes': {
                'distritos': distritos_ejemplo,
                'data': distritos_data,
                'meses': [mes['mes'].strftime('%Y-%m') for mes in meses]
            },
            'tipos_mes': {
                'tipos': tipos_ejemplo,
                'data': tipos_data,
                'meses': [mes['mes'].strftime('%Y-%m') for mes in meses]
            },
            'presupuesto_mes': presupuesto_data,
            'caracteristicas_mes': caracteristicas_data,
            'metricas': {
                'totales': totales,
                'crecimiento': crecimiento,
                'picos': picos,
                'valles': valles,
                'tendencia': tendencia,
            },
            'insights': insights,
            'is_sample_data': True,
            'message': 'Se están mostrando datos de ejemplo porque la base de datos está vacía.'
        }
        
        from django.http import JsonResponse
        return JsonResponse(response_data, safe=False)


class ApiAnalisisProgresoView(TemplateView):
    """API para consultar progreso de análisis asíncrono."""
    
    def get(self, request, task_id, *args, **kwargs):
        progreso = obtener_progreso_tarea(task_id)
        
        # Mapear estados para compatibilidad con frontend
        status_map = {
            'processing': 'PROGRESS',
            'completed': 'SUCCESS',
            'failed': 'FAILURE',
            'pending': 'PENDING',
            'unknown': 'PENDING'
        }
        
        status = progreso.get('status', 'unknown')
        mapped_status = status_map.get(status, 'PENDING')
        
        response_data = {
            'task_id': task_id,
            'status': mapped_status,
            'progress': progreso.get('progress', 0),
            'message': progreso.get('message', ''),
            'current_step': progreso.get('current_step', 'Procesando'),
            'generated_at': progreso.get('generated_at')
        }
        
        # Si la tarea está completada, incluir los datos en 'result'
        if mapped_status == 'SUCCESS' and 'data' in progreso:
            response_data['result'] = progreso['data']
        
        # Si hay error, incluir detalles
        if mapped_status == 'FAILURE':
            response_data['error'] = progreso.get('error', 'Error desconocido')
        
        return JsonResponse(response_data)


class ExportarAnalisisExcelView(TemplateView):
    """Exporta el análisis a Excel."""
    
    def get(self, request, *args, **kwargs):
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from django.http import HttpResponse
            import io
            
            # Crear libro de trabajo
            wb = Workbook()
            ws = wb.active
            ws.title = "Análisis Temporal"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título
            ws.merge_cells('A1:G1')
            title_cell = ws['A1']
            title_cell.value = "ANÁLISIS TEMPORAL DE REQUERIMIENTOS INMOBILIARIOS"
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = Alignment(horizontal="center")
            
            # Subtítulo con filtros aplicados
            ws.merge_cells('A2:G2')
            subtitle = f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')}"
            ws['A2'].value = subtitle
            ws['A2'].alignment = Alignment(horizontal="center")
            ws['A2'].font = Font(italic=True)
            
            # Encabezados de sección
            headers = ['Mes', 'Total Requerimientos', 'Compra', 'Alquiler',
                      'Departamento', 'Casa', 'Terreno', 'Presupuesto Promedio']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Obtener datos (similar a la API)
            fecha_inicio = request.GET.get('fecha_inicio')
            fecha_fin = request.GET.get('fecha_fin')
            filtros = {
                'condicion': request.GET.get('condicion'),
                'tipo_propiedad': request.GET.get('tipo_propiedad'),
                'distrito': request.GET.get('distrito'),
                'fuente': request.GET.get('fuente'),
            }
            
            datos_mes = list(obtener_requerimientos_por_mes(fecha_inicio, fecha_fin, filtros))
            
            # Llenar datos
            for row, dato in enumerate(datos_mes, 5):
                ws.cell(row=row, column=1, value=dato['mes'].strftime('%B %Y') if dato['mes'] else '')
                ws.cell(row=row, column=2, value=dato['total'])
                ws.cell(row=row, column=3, value=dato['compra'])
                ws.cell(row=row, column=4, value=dato['alquiler'])
                ws.cell(row=row, column=5, value=dato['departamento'])
                ws.cell(row=row, column=6, value=dato['casa'])
                ws.cell(row=row, column=7, value=dato['terreno'])
                ws.cell(row=row, column=8, value=float(dato['presupuesto_promedio']) if dato['presupuesto_promedio'] else 0)
            
            # Ajustar anchos de columna
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Preparar respuesta
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="analisis_temporal_requerimientos.xlsx"'
            return response
            
        except ImportError:
            from django.http import HttpResponse
            return HttpResponse("Error: openpyxl no está instalado. Ejecute 'pip install openpyxl'", status=500)
        except Exception as e:
            from django.http import HttpResponse
            return HttpResponse(f"Error al generar Excel: {str(e)}", status=500)


class ExportarAnalisisPDFView(TemplateView):
    """Exporta el análisis a PDF."""
    
    def get(self, request, *args, **kwargs):
        try:
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from django.http import HttpResponse
            import io
            
            # Crear buffer para PDF
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
            elements = []
            styles = getSampleStyleSheet()
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=12,
                alignment=1  # Centrado
            )
            title = Paragraph("ANÁLISIS TEMPORAL DE REQUERIMIENTOS INMOBILIARIOS", title_style)
            elements.append(title)
            
            # Subtítulo
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Normal'],
                fontSize=10,
                spaceAfter=24,
                alignment=1,
                textColor=colors.grey
            )
            subtitle = Paragraph(
                f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')} | "
                f"Sistema de Análisis Inmobiliario",
                subtitle_style
            )
            elements.append(subtitle)
            
            # Obtener datos
            fecha_inicio = request.GET.get('fecha_inicio')
            fecha_fin = request.GET.get('fecha_fin')
            filtros = {
                'condicion': request.GET.get('condicion'),
                'tipo_propiedad': request.GET.get('tipo_propiedad'),
                'distrito': request.GET.get('distrito'),
                'fuente': request.GET.get('fuente'),
            }
            
            datos_mes = list(obtener_requerimientos_por_mes(fecha_inicio, fecha_fin, filtros))
            
            # Crear tabla de datos
            if datos_mes:
                table_data = []
                # Encabezados
                headers = ['Mes', 'Total', 'Compra', 'Alquiler', 'Depto.', 'Casa', 'Terreno', 'Presupuesto']
                table_data.append(headers)
                
                # Filas de datos
                for dato in datos_mes:
                    row = [
                        dato['mes'].strftime('%b %Y') if dato['mes'] else '',
                        str(dato['total']),
                        str(dato['compra']),
                        str(dato['alquiler']),
                        str(dato['departamento']),
                        str(dato['casa']),
                        str(dato['terreno']),
                        f"${dato['presupuesto_promedio']:,.0f}" if dato['presupuesto_promedio'] else '$0'
                    ]
                    table_data.append(row)
                
                # Crear tabla
                table = Table(table_data, colWidths=[1.2*inch, 0.7*inch, 0.7*inch, 0.7*inch,
                                                     0.7*inch, 0.7*inch, 0.7*inch, 1*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ]))
                
                elements.append(Spacer(1, 0.25*inch))
                elements.append(table)
            
            # Nota al pie
            elements.append(Spacer(1, 0.5*inch))
            note_style = ParagraphStyle(
                'NoteStyle',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.grey,
                alignment=1
            )
            note = Paragraph(
                "Este reporte fue generado automáticamente por el Sistema de Análisis Temporal. "
                "Los datos están sujetos a actualizaciones periódicas.",
                note_style
            )
            elements.append(note)
            
            # Construir PDF
            doc.build(elements)
            
            # Preparar respuesta
            buffer.seek(0)
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="analisis_temporal_requerimientos.pdf"'
            return response
            
        except ImportError:
            from django.http import HttpResponse
            return HttpResponse("Error: reportlab no está instalado. Ejecute 'pip install reportlab'", status=500)
        except Exception as e:
            from django.http import HttpResponse
            return HttpResponse(f"Error al generar PDF: {str(e)}", status=500)


# ── Helper: normalizar teléfono ─────────────────────────────────
def _normalizar_telefono(valor):
    """Extrae solo dígitos de un teléfono para comparación flexible.
    
    '999888777' → '999888777'
    '+51999888777' → '51999888777'
    '999 888 777' → '999888777'
    """
    import re
    return re.sub(r'\D', '', valor)


def _buscar_agente_por_telefono(telefono_raw):
    """Busca un Agente cuyo teléfono coincida con los dígitos del número dado.
    
    Primero intenta coincidencia exacta, luego busca por dígitos normalizados
    usando LIKE con % para ignorar prefijos como +51.
    """
    from agentes.models import Agente
    
    digitos = _normalizar_telefono(telefono_raw)
    if not digitos or len(digitos) < 6:
        return None
    
    # 1. Coincidencia exacta
    agente = Agente.objects.filter(telefono=telefono_raw.strip()).first()
    if agente:
        return agente
    
    # 2. Coincidencia por dígitos (el teléfono en BD contiene los mismos dígitos)
    #    Ej: BD tiene "+51999888777", buscamos "999888777" → LIKE '%999888777%'
    from django.db.models import Q
    agente = Agente.objects.filter(
        Q(telefono__endswith=digitos) |
        Q(telefono__contains=digitos)
    ).first()
    if agente:
        return agente
    
    # 3. Último intento: extraer solo dígitos del campo telefono en BD y comparar
    #    Esto captura casos donde el formato es muy diferente
    todos = Agente.objects.all()
    for a in todos:
        if _normalizar_telefono(a.telefono) == digitos:
            return a
    
    return None


def _buscar_agente_por_nombre(nombre_raw):
    """Busca un Agente cuyo nombre coincida parcialmente (insensible a mayúsculas).
    
    Útil cuando el teléfono no coincide pero el nombre es muy similar.
    Ej: 'Monica Valdivia Ponce' ≈ 'Monica Valdivia'
    
    Estrategia de búsqueda (en orden):
    1. Coincidencia exacta insensible (iexact)
    2. El nombre buscado está CONTENIDO en algún nombre de BD (icontains)
    3. Algún nombre de BD está CONTENIDO en el nombre buscado (icontains inverso)
    4. Todas las palabras del nombre buscado (≥3 letras) aparecen en un nombre de BD
    5. Primer nombre + primer apellido coinciden
    """
    from agentes.models import Agente
    from django.db.models import Q
    
    nombre = nombre_raw.strip().lower()
    if not nombre or len(nombre) < 3:
        return None
    
    # 1. Coincidencia exacta (insensible)
    agente = Agente.objects.filter(nombre_completo__iexact=nombre_raw.strip()).first()
    if agente:
        return agente
    
    # 2. El nombre buscado está contenido en algún nombre de BD
    #    Ej: buscamos "Keren" y en BD hay "Keren Aragon"
    agente = Agente.objects.filter(nombre_completo__icontains=nombre_raw.strip()).first()
    if agente:
        return agente
    
    # 3. Algún nombre de BD está contenido en el nombre buscado
    #    Ej: buscamos "Keren Aragon Paredes" y en BD hay "Keren Aragon"
    todos = Agente.objects.all()
    for a in todos:
        if a.nombre_completo and a.nombre_completo.strip().lower() in nombre:
            return a
    
    # 4. Todas las palabras del nombre buscado (≥3 letras) aparecen en un nombre de BD
    palabras = [p for p in nombre.split() if len(p) >= 3]
    if len(palabras) >= 2:
        q = Q(nombre_completo__icontains=palabras[0])
        for palabra in palabras[1:]:
            q &= Q(nombre_completo__icontains=palabra)
        agente = Agente.objects.filter(q).first()
        if agente:
            return agente
    
    # 5. Búsqueda por palabra más significativa (primer nombre + apellido)
    if len(palabras) >= 2:
        primer_nombre = palabras[0]
        primer_apellido = palabras[1]
        if len(primer_nombre) >= 3 and len(primer_apellido) >= 3:
            agente = Agente.objects.filter(
                Q(nombre_completo__icontains=primer_nombre) &
                Q(nombre_completo__icontains=primer_apellido)
            ).first()
            if agente:
                return agente
    
    return None


# ── API: Buscar agente por teléfono ─────────────────────────────
class BuscarAgentePorTelefonoView(View):
    """Busca un agente en la tabla Agente por número de teléfono.

    GET /requerimientos/api/buscar-agente/?telefono=999888777
    Retorna JSON con datos del agente si existe, o encontrado=false.
    La búsqueda normaliza dígitos para ignorar diferencias de formato.
    """
    def get(self, request, *args, **kwargs):
        telefono = request.GET.get('telefono', '').strip()
        if not telefono or len(telefono) < 6:
            return JsonResponse({'encontrado': False})
        try:
            agente = _buscar_agente_por_telefono(telefono)
            if agente:
                return JsonResponse({
                    'encontrado': True,
                    'id': agente.id,
                    'nombre_completo': agente.nombre_completo,
                    'tipo_agente': agente.tipo_agente,
                    'telefono': agente.telefono or '',
                })
            return JsonResponse({'encontrado': False})
        except Exception as e:
            return JsonResponse({'encontrado': False, 'error': str(e)})


# ── API: Buscar agente por nombre ───────────────────────────────
class BuscarAgentePorNombreView(View):
    """Busca un agente en la tabla Agente por nombre.

    GET /requerimientos/api/buscar-agente-por-nombre/?nombre=Juan%20Perez
    Retorna JSON con datos del agente si existe, o encontrado=false.
    Útil cuando el requerimiento tiene nombre de agente pero no teléfono.
    """
    def get(self, request, *args, **kwargs):
        nombre = request.GET.get('nombre', '').strip()
        if not nombre or len(nombre) < 3:
            return JsonResponse({'encontrado': False})
        try:
            agente = _buscar_agente_por_nombre(nombre)
            if agente:
                return JsonResponse({
                    'encontrado': True,
                    'id': agente.id,
                    'nombre_completo': agente.nombre_completo,
                    'tipo_agente': agente.tipo_agente,
                    'telefono': agente.telefono or '',
                })
            return JsonResponse({'encontrado': False})
        except Exception as e:
            return JsonResponse({'encontrado': False, 'error': str(e)})


class CrearAgenteRapidoView(View):
    """Crea un agente rápido desde el modal de requerimientos.

    POST /requerimientos/api/crear-agente-rapido/
    Body JSON: { telefono: "...", nombre_completo: "..." }
    Idempotente: busca primero por teléfono (dígitos normalizados) y
    también por nombre similar para evitar duplicados aunque el número
    tenga dígitos diferentes (error de captura).

    El teléfono se normaliza (sin espacios, guiones) antes de guardar
    para mantener consistencia en la base de datos.

    Si el requerimiento no tiene teléfono (solo nombre), se permite
    crear el agente igualmente con el nombre proporcionado.
    """
    def post(self, request, *args, **kwargs):
        try:
            import json
            import re
            body = json.loads(request.body)
            telefono_raw = body.get('telefono', '').strip()
            nombre_completo = body.get('nombre_completo', '').strip()

            from agentes.models import Agente

            # ── Validación: se requiere al menos nombre o teléfono ──
            if not nombre_completo and (not telefono_raw or len(telefono_raw) < 6):
                return JsonResponse({'ok': False, 'error': 'Se requiere nombre o teléfono del agente'})

            # ── Normalizar teléfono si existe ──
            telefono_normalizado = ''
            if telefono_raw and len(telefono_raw) >= 6:
                # '+51 958 190 107' → '+51958190107'
                telefono_normalizado = re.sub(r'[\s\-\(\)]+', '', telefono_raw)

            # ── Buscar existente por teléfono ──
            if telefono_normalizado:
                agente = _buscar_agente_por_telefono(telefono_normalizado)
                if agente:
                    # Actualizar teléfono si cambió el formato
                    if agente.telefono != telefono_normalizado:
                        agente.telefono = telefono_normalizado
                        agente.save()
                    # Actualizar nombre si se proporcionó uno mejor
                    if nombre_completo and agente.nombre_completo != nombre_completo:
                        agente.nombre_completo = nombre_completo
                        agente.save()
                    return JsonResponse({
                        'ok': True,
                        'id': agente.id,
                        'nombre_completo': agente.nombre_completo,
                        'telefono': agente.telefono or '',
                    })

            # ── Buscar existente por nombre ──
            if nombre_completo:
                agente = _buscar_agente_por_nombre(nombre_completo)
                if agente:
                    # Actualizar teléfono si se proporcionó uno y el agente no tenía
                    if telefono_normalizado and not agente.telefono:
                        agente.telefono = telefono_normalizado
                        agente.save()
                    elif telefono_normalizado and agente.telefono != telefono_normalizado:
                        agente.telefono = telefono_normalizado
                        agente.save()
                    return JsonResponse({
                        'ok': True,
                        'id': agente.id,
                        'nombre_completo': agente.nombre_completo,
                        'telefono': agente.telefono or '',
                    })

            # ── Anti-duplicado: búsqueda final ultra-agresiva antes de crear ──
            # Si hay nombre, buscar en TODOS los agentes por coincidencia parcial
            # para evitar crear duplicados aunque _buscar_agente_por_nombre no haya
            # encontrado (por diferencias mínimas como tildes, espacios, etc.)
            if nombre_completo:
                from django.db.models import Q as Q2
                nombre_busqueda = nombre_completo.strip().lower()
                palabras_busqueda = [p for p in nombre_busqueda.split() if len(p) >= 3]
                if len(palabras_busqueda) >= 2:
                    # Buscar cualquier agente que contenga al menos 2 palabras del nombre
                    q_dup = Q2(nombre_completo__icontains=palabras_busqueda[0])
                    for p in palabras_busqueda[1:]:
                        q_dup |= Q2(nombre_completo__icontains=p)
                    dup = Agente.objects.filter(q_dup).first()
                    if dup:
                        # Actualizar teléfono si se proporcionó
                        if telefono_normalizado and not dup.telefono:
                            dup.telefono = telefono_normalizado
                            dup.save()
                        elif telefono_normalizado and dup.telefono != telefono_normalizado:
                            dup.telefono = telefono_normalizado
                            dup.save()
                        return JsonResponse({
                            'ok': True,
                            'id': dup.id,
                            'nombre_completo': dup.nombre_completo,
                            'telefono': dup.telefono or '',
                            'actualizado': True,
                        })

            # ── No existe: crear nuevo ──
            agente = Agente.objects.create(
                telefono=telefono_normalizado,
                nombre_completo=nombre_completo or telefono_normalizado,
                tipo_agente=Agente.TipoAgente.INDEPENDIENTE,
            )

            return JsonResponse({
                'ok': True,
                'id': agente.id,
                'nombre_completo': agente.nombre_completo,
                'telefono': agente.telefono or '',
            })
        except Exception as e:
            return JsonResponse({'ok': False, 'error': str(e)})


class ApiZonaCalleAutocompleteView(View):
    """API de autocomplete para ZonaCalle (zonas y calles)."""
    def get(self, request, *args, **kwargs):
        q = request.GET.get('q', '').strip()
        if len(q) < 1:
            return JsonResponse({'data': []})
        resultados = ZonaCalle.objects.filter(
            nombre__icontains=q
        ).order_by('-veces_usado', 'nombre')[:20]
        data = [{
            'id': zc.id,
            'nombre': zc.nombre,
            'veces_usado': zc.veces_usado,
        } for zc in resultados]
        return JsonResponse({'data': data})


class ConfiguracionCalidadView(TemplateView):
    """Vista del dashboard de configuración de Quality Score."""
    template_name = 'requerimientos/config_calidad.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        config_obj = ConfiguracionCalidad.objects.filter(activo=True).first()
        if config_obj:
            ctx['config_obj'] = config_obj
            ctx['config'] = config_obj.config
        else:
            ctx['config_obj'] = None
            from .models import CONFIG_CALIDAD_DEFAULT
            ctx['config'] = CONFIG_CALIDAD_DEFAULT
        return ctx


@method_decorator(csrf_exempt, name='dispatch')
class ApiConfiguracionCalidadView(View):
    """API para guardar la configuración de Quality Score."""

    def get(self, request, *args, **kwargs):
        """Retorna la configuración activa."""
        config_obj = ConfiguracionCalidad.objects.filter(activo=True).first()
        if config_obj:
            return JsonResponse({'success': True, 'config': config_obj.config, 'id': config_obj.id})
        from .models import CONFIG_CALIDAD_DEFAULT
        return JsonResponse({'success': True, 'config': CONFIG_CALIDAD_DEFAULT, 'id': None})

    def post(self, request, *args, **kwargs):
        """Guarda la configuración."""
        try:
            body = json.loads(request.body)
            config_data = body.get('config', {})

            # Desactivar configuraciones anteriores
            ConfiguracionCalidad.objects.filter(activo=True).update(activo=False)

            # Crear nueva configuración
            nombre = body.get('nombre', f'Config {timezone.now().strftime("%Y-%m-%d %H:%M")}')
            config_obj = ConfiguracionCalidad.objects.create(
                activo=True,
                config=config_data,
                nombre=nombre,
            )
            return JsonResponse({'success': True, 'id': config_obj.id})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class ApiEstadisticasCalidadView(View):
    """API para obtener estadísticas de calidad de todos los requerimientos."""

    def get(self, request, *args, **kwargs):
        from .analytics import calcular_quality_score
        from .models import CONFIG_CALIDAD_DEFAULT

        config_obj = ConfiguracionCalidad.objects.filter(activo=True).first()
        config = config_obj.config if config_obj else CONFIG_CALIDAD_DEFAULT

        requerimientos = Requerimiento.objects.all()
        total = requerimientos.count()

        if total == 0:
            return JsonResponse({
                'success': True,
                'total': 0,
                'promedio': 0,
                'niveles': {'Excelente': 0, 'Bueno': 0, 'Regular': 0, 'Malo': 0},
                'dimensiones_promedio': {
                    'completitud': 0, 'especificidad': 0,
                    'presupuesto': 0, 'antiguedad': 0,
                },
                'distribucion_scores': [],
            })

        scores = []
        niveles_count = {'Excelente': 0, 'Bueno': 0, 'Regular': 0, 'Malo': 0}
        dim_sum = {'completitud': 0.0, 'especificidad': 0.0, 'presupuesto': 0.0, 'antiguedad': 0.0}

        for req in requerimientos.iterator(chunk_size=500):
            qs = calcular_quality_score(req, config)
            scores.append({
                'id': req.id,
                'score': qs['score'],
                'nivel': qs['nivel'],
                'dimensiones': qs['dimensiones'],
            })
            niveles_count[qs['nivel']] = niveles_count.get(qs['nivel'], 0) + 1
            for k in dim_sum:
                dim_sum[k] += qs['dimensiones'].get(k, 0)

        promedio = sum(s['score'] for s in scores) / total if total else 0

        # Distribución de scores (agrupada en rangos de 10)
        distribucion = {}
        for s in scores:
            rango = int(s['score'] // 10) * 10
            key = f'{rango}-{rango+9}'
            distribucion[key] = distribucion.get(key, 0) + 1

        return JsonResponse({
            'success': True,
            'total': total,
            'promedio': round(promedio, 1),
            'niveles': niveles_count,
            'dimensiones_promedio': {
                k: round(v / total, 1) for k, v in dim_sum.items()
            },
            'distribucion_scores': [
                {'rango': k, 'count': v}
                for k, v in sorted(distribucion.items(), key=lambda x: int(x[0].split('-')[0]))
            ],
        })

class ApiRecalcularQualityView(View):
    """API para recalcular y persistir quality scores de todos los requerimientos."""

    def post(self, request, *args, **kwargs):
        from .analytics import _get_calidad_config
        config = _get_calidad_config()
        requerimientos = Requerimiento.objects.all()
        total = requerimientos.count()
        actualizados = 0
        errores = 0

        for req in requerimientos.iterator(chunk_size=500):
            try:
                req.recalcular_quality_score(config=config, guardar=True)
                actualizados += 1
            except Exception:
                errores += 1

        return JsonResponse({
            'success': True,
            'total': total,
            'actualizados': actualizados,
            'errores': errores,
            'mensaje': f'Se recalcularon {actualizados} de {total} requerimientos.'
        })


class ClonarRequerimientoView(View):
    """Vista para clonar un requerimiento (crear uno nuevo a partir de otro) vía POST (AJAX)."""
    CAMPOS_CLONABLES = [
        'verificado', 'fuente', 'agente', 'agente_telefono',
        'condicion', 'tipo_propiedad', 'distritos',
        'urbanizacion', 'zona',
        'presupuesto_monto', 'presupuesto_moneda', 'presupuesto_forma_pago',
        'habitaciones', 'banos', 'cochera', 'ascensor', 'amueblado',
        'area_m2', 'piso_preferencia', 'caracteristicas_extra',
        'tipo_original', 'requerimiento',
    ]

    def post(self, request, *args, **kwargs):
        import json
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, TypeError):
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        # Crear nuevo requerimiento con los datos recibidos
        nuevo_req = Requerimiento()

        for campo in self.CAMPOS_CLONABLES:
            if campo in data:
                valor = data[campo]
                if campo in ('presupuesto_monto', 'area_m2', 'habitaciones', 'banos'):
                    if valor == '' or valor is None:
                        setattr(nuevo_req, campo, None)
                    else:
                        try:
                            valor_convertido = float(valor) if campo == 'presupuesto_monto' else int(valor)
                            setattr(nuevo_req, campo, valor_convertido)
                        except (ValueError, TypeError):
                            setattr(nuevo_req, campo, None)
                else:
                    setattr(nuevo_req, campo, valor)

        # ── Sincronizar ZonaCalle con los tags del campo 'zona' ──
        if 'zona' in data and data['zona']:
            tags = [t.strip() for t in data['zona'].split(',') if t.strip()]
            for tag in tags:
                obj, created = ZonaCalle.objects.get_or_create(
                    nombre__iexact=tag,
                    defaults={'nombre': tag, 'veces_usado': 1}
                )
                if not created:
                    if obj.nombre != tag:
                        obj.nombre = tag
                    ZonaCalle.objects.filter(pk=obj.pk).update(veces_usado=models.F('veces_usado') + 1)

        nuevo_req.save()

        return JsonResponse({
            'ok': True,
            'pk': nuevo_req.pk,
            'mensaje': f'Requerimiento clonado correctamente (ID: {nuevo_req.pk})',
        })
