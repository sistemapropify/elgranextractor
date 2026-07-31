"""
Exporta todos los registros de la tabla property a un Excel,
resolviendo campos indexados (FKs) a sus nombres legibles.
"""
import os, sys
from datetime import datetime
from decimal import Decimal
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'settings')
sys.path.insert(0, r'd:\PROMETEO\webapp')
import django; django.setup()
from django.db import connections
import openpyxl
from openpyxl.styles import Font, PatternFill

conn = connections['propifai']
cur = conn.cursor()

# Obtener datos de property
cur.execute("SELECT * FROM property ORDER BY id")
cols = [desc[0] for desc in cur.description]
props = cur.fetchall()
col_idx = {name: i for i, name in enumerate(cols)}
print(f"Propiedades: {len(props)}, Columnas: {len(cols)}")

def load_map(sql, name_len=80):
    """Carga {id: label} desde un query SELECT id, label."""
    try:
        cur.execute(sql)
        return {r[0]: str(r[1] or '')[:name_len] if r[1] is not None else f"ID:{r[0]}" for r in cur.fetchall()}
    except Exception as e:
        print(f"  ERROR: {sql} -> {e}")
        return {}

# --- Resolver cada FK ---
resolved = {}

# contact (agentes/contactos)
resolved['contact_id'] = load_map("SELECT id, CASE WHEN first_name IS NULL OR first_name='' THEN '' ELSE first_name + ' ' + ISNULL(last_name,'') END FROM contact")
resolved['created_by_id'] = resolved['contact_id']
resolved['updated_by_id'] = resolved['contact_id']
resolved['responsible_id'] = resolved['contact_id']
print(f"  contact_id/created_by_id/updated_by_id/responsible_id -> contact ({len(resolved['contact_id'])})")

# currency
resolved['currency_id'] = load_map("SELECT id, code FROM currency")
print(f"  currency_id -> currency.code ({len(resolved['currency_id'])})")

# district
resolved['district_id'] = load_map("SELECT id, name FROM district")
print(f"  district_id -> district.name ({len(resolved['district_id'])})")

# operation_type
resolved['operation_type_id'] = load_map("SELECT id, name FROM operation_type")
print(f"  operation_type_id -> operation_type.name ({len(resolved['operation_type_id'])})")

# payment_method
resolved['payment_method_id'] = load_map("SELECT id, name FROM payment_method")
print(f"  payment_method_id -> payment_method.name ({len(resolved['payment_method_id'])})")

# property_condition
resolved['property_condition_id'] = load_map("SELECT id, name FROM property_condition")
print(f"  property_condition_id -> property_condition.name ({len(resolved['property_condition_id'])})")

# property_status
resolved['property_status_id'] = load_map("SELECT id, name FROM property_status")
print(f"  property_status_id -> property_status.name ({len(resolved['property_status_id'])})")

# property_subtype
resolved['property_subtype_id'] = load_map("SELECT id, name FROM property_subtype")
print(f"  property_subtype_id -> property_subtype.name ({len(resolved['property_subtype_id'])})")

# property_type
resolved['property_type_id'] = load_map("SELECT id, name FROM property_type")
print(f"  property_type_id -> property_type.name ({len(resolved['property_type_id'])})")

# urbanization
resolved['urbanization_id'] = load_map("SELECT id, name FROM urbanization")
print(f"  urbanization_id -> urbanization.name ({len(resolved['urbanization_id'])})")

# parent_project_id -> self-reference a property.title
resolved['parent_project_id'] = load_map("SELECT id, title FROM property")
print(f"  parent_project_id -> property.title ({len(resolved['parent_project_id'])})")

# typology
resolved['typology_id'] = load_map("SELECT id, name FROM typology")
print(f"  typology_id -> typology.name ({len(resolved['typology_id'])})")

# wp_post_id -> es un ID de WordPress, no FK de BD
print(f"  wp_post_id -> (ID externo de WordPress, sin resolver)")

# Crear Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Propiedades"

headers = []
for c in cols:
    headers.append(c)
    if c in resolved:
        headers.append(f"{c}_nombre")

hf = Font(bold=True, color="FFFFFF")
hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for ci, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.font = hf; cell.fill = hfill

fmt = lambda dt: dt.strftime('%Y-%m-%d %H:%M') if hasattr(dt, 'strftime') and dt else ''
y_n = lambda v: 'Si' if v else 'No'

for r in range(len(props)):
    row_data = []
    for c in cols:
        val = props[r][col_idx[c]]
        if val is None:
            val = ''
        elif hasattr(val, 'strftime'):
            val = fmt(val)
        elif isinstance(val, bool):
            val = y_n(val)
        row_data.append(val)
        
        if c in resolved:
            id_val = props[r][col_idx[c]]
            if id_val is not None:
                row_data.append(resolved[c].get(id_val, ''))
            else:
                row_data.append('')
    
    for ci, val in enumerate(row_data, 1):
        ws.cell(row=r+2, column=ci, value=val)

for ci in range(1, len(headers)+1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(15, len(str(headers[ci-1]))+3)

out = r'd:\PROMETEO\propiedades_completas.xlsx'
wb.save(out)
print(f"\nOK {out}")
print(f"  {len(props)} registros | {len(headers)} columnas")
